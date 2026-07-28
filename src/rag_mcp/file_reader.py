"""
File reader with encoding detection, binary detection, and PDF extraction.

Handles reading text files with multiple encoding fallbacks and extracting
text from PDF documents for the RAG indexing pipeline.
"""

import hashlib
import sys
from dataclasses import dataclass
from pathlib import Path

from rag_mcp.pdf_extractor import PDFExtractor
from rag_mcp.pdf_converter import PDFConverter


@dataclass
class FileContent:
    """Represents the content of a successfully read file."""

    path: Path
    relative_path: str
    content: str
    file_hash: str
    is_pdf: bool = False


class FileReader:
    """Reads text and PDF files with encoding fallback and binary detection."""

    ENCODINGS = ["utf-8", "latin-1", "cp1252"]
    BINARY_CHECK_SIZE = 8192

    def __init__(self, pdf_cache_dir: Path | None = None):
        self.pdf_extractor = PDFExtractor()
        self.pdf_converter = PDFConverter()
        # Directory for converted PDF markdown when writing next to the source
        # isn't possible (e.g. read-only mounts). Set from config by the pipeline.
        self.pdf_cache_dir = pdf_cache_dir

    def _pdf_cache_path(self, pdf_path: Path) -> Path | None:
        """Deterministic cache path for a PDF's converted markdown, or None."""
        if not self.pdf_cache_dir:
            return None
        digest = hashlib.md5(str(pdf_path.resolve()).encode("utf-8")).hexdigest()[:12]
        return Path(self.pdf_cache_dir) / f"{pdf_path.stem}_{digest}.md"

    def read(self, filepath: Path, base_path: Path) -> FileContent | None:
        """
        Read a file and return its content.
        Returns None if file is binary, unreadable, or empty PDF.
        """
        relative_path = str(filepath.relative_to(base_path)).replace("\\", "/")
        file_hash = self.compute_hash(filepath)

        # PDF files — always convert to Markdown (never index raw PDF bytes).
        if filepath.suffix.lower() == ".pdf":
            # 1) Prefer an up-to-date .md sibling (e.g. produced by --convert-pdfs
            #    in writable setups, or a hand-authored companion doc).
            md_path = filepath.with_suffix(".md")
            if md_path.exists() and md_path.stat().st_mtime >= filepath.stat().st_mtime:
                content = self.read_text(md_path)
                if content:
                    return FileContent(
                        path=md_path,
                        relative_path=str(md_path.relative_to(base_path)).replace("\\", "/"),
                        content=content,
                        file_hash=file_hash,
                        is_pdf=True,
                    )

            # 2) Convert to the writable cache (works on read-only source mounts).
            cache_md = self._pdf_cache_path(filepath)
            if cache_md is not None:
                if not (cache_md.exists() and cache_md.stat().st_mtime >= filepath.stat().st_mtime):
                    self.pdf_converter.convert(filepath, output_path=cache_md)
                if cache_md.exists():
                    content = self.read_text(cache_md)
                    if content:
                        # Report the logical PDF path so search results point at
                        # the real document (not the cache file).
                        return FileContent(
                            path=cache_md,
                            relative_path=relative_path,
                            content=content,
                            file_hash=file_hash,
                            is_pdf=True,
                        )
            else:
                # 3) No cache configured: convert alongside the source (legacy).
                converted_path = self.pdf_converter.convert(filepath)
                if converted_path and converted_path.exists():
                    content = self.read_text(converted_path)
                    if content:
                        return FileContent(
                            path=converted_path,
                            relative_path=str(converted_path.relative_to(base_path)).replace("\\", "/"),
                            content=content,
                            file_hash=file_hash,
                            is_pdf=True,
                        )

            # 4) Last resort (image-only / unconvertible PDF): direct extraction.
            content = self.pdf_extractor.extract(filepath)
            if content is None:
                print(
                    f"  [warning] PDF has no extractable text: {relative_path}",
                    file=sys.stderr,
                )
                return None
            return FileContent(
                path=filepath,
                relative_path=relative_path,
                content=content,
                file_hash=file_hash,
                is_pdf=True,
            )

        # Excel files (.xlsx, .xls)
        if filepath.suffix.lower() in (".xlsx", ".xls"):
            content = self.read_excel(filepath)
            if content is None:
                print(
                    f"  [warning] Could not read Excel file: {relative_path}",
                    file=sys.stderr,
                )
                return None
            return FileContent(
                path=filepath,
                relative_path=relative_path,
                content=content,
                file_hash=file_hash,
                is_pdf=False,
            )

        # Binary check for non-PDF/non-Excel files
        if self.is_binary(filepath):
            return None

        # Text files
        content = self.read_text(filepath)
        if content is None:
            print(
                f"  [warning] Could not read file (encoding failure): {relative_path}",
                file=sys.stderr,
            )
            return None

        return FileContent(
            path=filepath,
            relative_path=relative_path,
            content=content,
            file_hash=file_hash,
            is_pdf=False,
        )

    def is_binary(self, filepath: Path) -> bool:
        """Check for null bytes in first 8192 bytes."""
        try:
            with open(filepath, "rb") as f:
                chunk = f.read(self.BINARY_CHECK_SIZE)
                return b"\x00" in chunk
        except (OSError, IOError):
            return True

    def compute_hash(self, filepath: Path) -> str:
        """Compute MD5 hash for change detection."""
        hasher = hashlib.md5()
        try:
            with open(filepath, "rb") as f:
                for block in iter(lambda: f.read(8192), b""):
                    hasher.update(block)
        except (OSError, IOError):
            return ""
        return hasher.hexdigest()

    def read_text(self, filepath: Path) -> str | None:
        """Try multiple encodings, return content or None."""
        for encoding in self.ENCODINGS:
            try:
                with open(filepath, "r", encoding=encoding) as f:
                    return f.read()
            except (UnicodeDecodeError, UnicodeError):
                continue
            except (OSError, IOError):
                return None
        return None

    def read_excel(self, filepath: Path) -> str | None:
        """
        Read an Excel file and convert to text (Markdown table format).
        Each sheet becomes a section with its data rendered as a table.
        Skips empty rows/columns and limits output to avoid excessive chunking.
        """
        try:
            from openpyxl import load_workbook

            wb = load_workbook(filepath, read_only=True, data_only=True)
            parts = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []

                for row in ws.iter_rows(values_only=True):
                    # Skip completely empty rows
                    if not any(cell is not None and str(cell).strip() for cell in row):
                        continue
                    rows.append(row)

                if not rows:
                    continue

                # Determine actual column count (trim trailing empty columns)
                max_col = 0
                for row in rows:
                    for i in range(len(row) - 1, -1, -1):
                        if row[i] is not None and str(row[i]).strip():
                            max_col = max(max_col, i + 1)
                            break

                if max_col == 0:
                    continue

                parts.append(f"## Sheet: {sheet_name}\n")

                # First row as header
                header = rows[0][:max_col]
                header_cells = [str(cell).strip() if cell is not None else "" for cell in header]
                parts.append("| " + " | ".join(header_cells) + " |")
                parts.append("| " + " | ".join(["---"] * max_col) + " |")

                # Data rows
                for row in rows[1:]:
                    cells = [str(cell).strip() if cell is not None else "" for cell in row[:max_col]]
                    # Pad if row is shorter
                    while len(cells) < max_col:
                        cells.append("")
                    parts.append("| " + " | ".join(cells) + " |")

                parts.append("")  # blank line between sheets

            wb.close()

            content = "\n".join(parts)
            return content if content.strip() else None

        except Exception as e:
            print(f"  [warning] Excel read error: {e}", file=sys.stderr)
            return None
